import openpyxl 
from openpyxl import load_workbook

arquivo=load_workbook('lista.xlsx')
planilha = arquivo.active

def cadastrar(cont):
        nome=input("\nNome: ")
        planilha['A'+str(cont)]= nome
        cpf=input("\nCPF: ")
        planilha['B'+str(cont)]= cpf
        telefone=input("\nTelefone:")
        planilha['C'+str(cont)]= telefone
        email=input("\nEmail: ")
        planilha['D'+str(cont)]= email
        endereco=input("\nEndereço: ")
        planilha['E'+str(cont)]= endereco
        data=input("\nData de nascimento: ")
        planilha['F'+str(cont)]= data
def atualizar():
    cont1=1
    encontrado=False
    edit=input("Digite o CPF do cliente: ")
    for linha in planilha.iter_rows(min_row=2, max_row=planilha.max_row, min_col=2, max_col=2):
        cont1+=1
        for celula in linha :
            if celula.value == edit:
                info=(planilha['A'+str(cont1)].value,planilha['B'+str(cont1)].value,planilha['C'+str(cont1)].value,planilha['D'+str(cont1)].value,planilha['E'+str(cont1)].value,planilha['F'+str(cont1)].value)
                print(info)
                print ("(1)Nome")
                print ("(2)CPF")
                print ("(3)Telefone")
                print ("(4)Email")
                print ("(5)Endereço")
                print ("(6)Data de nascimento")
                opcao=int(input())
                if opcao == 1:
                    nome=input("\nNome: ")
                    planilha['A'+str(cont1)]= nome
                if opcao == 2:    
                    cpf=input("\nCPF: ")
                    planilha['B'+str(cont1)]= cpf
                if opcao == 3:
                    telefone=input("\nTelefone:")
                    planilha['C'+str(cont1)]= telefone
                if opcao == 4:
                    email=input("\nEmail: ")
                    planilha['D'+str(cont1)]= email
                if opcao == 5:
                    endereco=input("\nEndereço: ")
                    planilha['E'+str(cont1)]= endereco
                if opcao == 6:
                    data=input("\nData de nascimento: ")   
                    planilha['F'+str(cont1)]= data
                        
                encontrado = True
    if not encontrado:
        print("Este CPF não existe !")
def listar():
    informacoes=[]
    for row in planilha.iter_rows(min_row=2, max_row=planilha.max_row, min_col=1, max_col=6):
        nome, cpf, telefone, email, address, birth = [cell.value for cell in row]
        print(f'Nome: {nome}, CPF: {cpf}, Telefone: {telefone}, E-mail: {email}, Endereço: {address}, Data de Nascimento: {birth}')
def excluir():
    encontrado= False
    delete=input("Digite o CPF do cliente que deseja remover: ")
    cont2=1
    for linha in planilha.iter_rows(min_row=2, max_row=planilha.max_row, min_col=2, max_col=2):
        cont2+=1
        for celula in linha :
            if celula.value == delete:
                info=(planilha['A'+str(cont2)].value,planilha['B'+str(cont2)].value,planilha['C'+str(cont2)].value,planilha['D'+str(cont2)].value,planilha['E'+str(cont2)].value,planilha['F'+str(cont2)].value)
                print(info)
                print ("Deseja excluir esse cliente? (1)Sim  (2)Não")
                escolha=int(input(""))
                if escolha==1:
                    planilha['A'+str(cont2)]= ""
                    planilha['B'+str(cont2)]= ""
                    planilha['C'+str(cont2)]= ""
                    planilha['D'+str(cont2)]= ""
                    planilha['E'+str(cont2)]= ""
                    planilha['F'+str(cont2)]= ""
                    arquivo.save('lista.xlsx')
                encontrado2 = True
    if not encontrado:
        print("Este CPF não existe !")
with open('save.txt', 'r') as arquivo2:
    conteudo = arquivo2.read()
cont=int(conteudo)
while(True):
    print(cont)
    print("O que deseja fazer?\n")
    print("(1)Cadastrar cliente")
    print("(2)Atualizar dados de cliente")
    print("(3)listar")
    print("(4)Excluir registros\n")
    opcao=int(input())
    match (opcao):
        case 1:
            cont+=1
            cadastrar(cont)
        case 2:
            atualizar()
        case 3:
            listar()
        case 4:
            excluir()
    with open('save.txt', 'w') as arquivo2:
        arquivo2.write(str(cont))
    arquivo.save('lista.xlsx')
    